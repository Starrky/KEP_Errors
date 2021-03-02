import glob
import shutil
import openpyxl
from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join
import pandas as pd
import csv
from time import time
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import os

start_time = time()
cwd = os.getcwd()
results_dir = f'{cwd}\\Results'
results_dir_csv = f'{results_dir}\\csv'
results_dir_xlsx = f'{results_dir}\\xlsx'

names_csv = [f for f in listdir(results_dir_csv) if isfile(join(results_dir_csv, f))]
for file in names_csv:
    loc = f"{results_dir_csv}\\{file}"
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(loc) as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            ws.append(row)

    file = file.removesuffix(".csv")
    wb.save(f'{results_dir_xlsx}/{file}.xlsx')

names_xlsx = [f for f in listdir(results_dir_xlsx) if isfile(join(results_dir_xlsx, f))]
dataframes = []

for file in names_xlsx:
    file = f"{results_dir_xlsx}\\{file}"
    wb = load_workbook(file)
    ws = wb.active
    ws.delete_rows(2, amount=1)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            User = ws[f'B{cell.row}'].value
            User = str(User)
            if 'sa' in User:
                ws[f'B{cell.row}'] = "sa"

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            Value = ws[f'A{cell.row}'].value
            Value = str(Value)

            if 'Sqlcmd: Error:' in Value or 'Msg 208' in Value or 'Invalid object name' in Value:
                ws.delete_rows(cell.row, amount=1)

            elif "rows" in Value:
                ws.delete_rows(cell.row, amount=1)
    wb.save(file)

for file in names_xlsx:
    xlsx = f"{results_dir_xlsx}\\{file}"
    df = pd.read_excel(xlsx)
    file = file.removesuffix(".xlsx")
    df.insert(0, 'Store', file)

    if df is not None:
        dataframes.append(df)

Final_excel = f'{results_dir}\\Errors_status.xlsx'
dataframes_df = pd.concat(dataframes)
dataframes_df.to_excel(Final_excel, index=False)

wb = load_workbook(Final_excel)
ws = wb.active
ws.delete_rows(1, amount=2)

for row in ws.iter_rows(min_row=2):
    # values = []
    # columns = 'ABCDEFGHIJ'
    # for column in columns:
    #     print(ws[f'{column}{cell.row}'].value)
    #     values.append(ws[f'{column}{cell.row}'].value.strip())

    for cell in row:
        Value = ws[f'A{cell.row}'].value
        Value = str(Value)

        if '-' in Value or '-' in Value:
            ws.delete_rows(cell.row, amount=1)

        elif "rows" in Value:
            ws.delete_rows(cell.row, amount=1)

        elif Value == 'None':
            ws.delete_rows(cell.row, amount=1)

        columns = 'ABCDEFGHIJ'
        for column in columns:
            value = ws[f'{column}{cell.row}'].value
            ws[f'{column}{cell.row}'].value = str(value).strip()


wb.save(Final_excel)

df = pd.read_excel(Final_excel)
df = df.fillna(0)
workbook = xlsxwriter.Workbook(Final_excel)
max_row = len(df) + 1
worksheet1 = workbook.add_worksheet('Tabela')

worksheet1.add_table(f'A1:J{max_row}', {'data': df.values.tolist(),
                                        'style': 'Table Style Light 9', 'columns':
                                            [
                                                {'header': 'Store'},
                                                {'header': 'ErrorID'},
                                                {'header': 'UserName'},
                                                {'header': 'ErrorNumber'},
                                                {'header': 'ErrorState'},
                                                {'header': 'ErrorSeverity'},
                                                {'header': 'ErrorLine'},
                                                {'header': 'ErrorProcedure'},
                                                {'header': 'ErrorMessage'},
                                                {'header': 'ErrorDateTime'}

                                            ]})

format1 = workbook.add_format({'bold': False, 'align': 'center'})

format1.set_align('Center')
worksheet1.set_row(0, 16, format1)
worksheet1.set_column('A:A', 18, format1)
worksheet1.set_column('B:B', 12, format1)
worksheet1.set_column('C:C', 14, format1)
worksheet1.set_column('D:D', 18, format1)
worksheet1.set_column('E:E', 18, format1)
worksheet1.set_column('F:F', 12, format1)
worksheet1.set_column('G:G', 18, format1)
worksheet1.set_column('H:H', 18, format1)
worksheet1.set_column('I:I', 70, format1)
worksheet1.set_column('J:J', 18, format1)
worksheet1.set_column('K:K', 18, format1)

workbook.close()


print("Process finished --- %s seconds ---" % (time() - start_time))
